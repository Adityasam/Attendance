import io
import json
import math
import calendar
from collections import OrderedDict
from datetime import datetime, date, timedelta
from flask import Blueprint, request, jsonify, session, send_file
from routes.db_helper import query, execute
from routes.auth import login_required

attendance_bp = Blueprint('attendance', __name__)


@attendance_bp.route('/', methods=['GET'])
@login_required
def list_attendance():
    date_filter = request.args.get('date', date.today().isoformat())
    project_id = request.args.get('project_id')
    month = request.args.get('month')
    year = request.args.get('year')

    if month and year:
        rows = query(
            '''SELECT a.*, e.FullName, e.Designation, e.Picture
               FROM main_attendance a
               LEFT JOIN main_employee e ON a.EmployeeCode = e.EmployeeCode
               WHERE strftime('%m', a.AttendanceDateTime) = ?
               AND strftime('%Y', a.AttendanceDateTime) = ?
               ORDER BY a.AttendanceDateTime DESC''',
            (month.zfill(2), str(year))
        )
    elif project_id:
        rows = query(
            '''SELECT a.*, e.FullName, e.Designation, e.Picture
               FROM main_attendance a
               LEFT JOIN main_employee e ON a.EmployeeCode = e.EmployeeCode
               WHERE date(a.AttendanceDateTime) = ? AND a.ProjectId = ?
               ORDER BY a.AttendanceDateTime DESC''',
            (date_filter, project_id)
        )
    else:
        rows = query(
            '''SELECT a.*, e.FullName, e.Designation, e.Picture
               FROM main_attendance a
               LEFT JOIN main_employee e ON a.EmployeeCode = e.EmployeeCode
               WHERE date(a.AttendanceDateTime) = ?
               ORDER BY a.AttendanceDateTime DESC''',
            (date_filter,)
        )
    return jsonify([dict(r) for r in rows])


@attendance_bp.route('/summary', methods=['GET'])
@login_required
def summary():
    """Dashboard summary: today present/absent counts."""
    today = date.today().isoformat()
    total_emp = query('SELECT COUNT(*) as cnt FROM main_employee WHERE Active=1 AND Deleted=0', one=True)['cnt']
    present_today = query(
        '''SELECT COUNT(DISTINCT EmployeeCode) as cnt FROM main_attendance
           WHERE date(AttendanceDateTime) = ?''', (today,), one=True
    )['cnt']

    current_month = date.today().strftime('%Y-%m')
    monthly_data = query(
        '''SELECT strftime('%d', AttendanceDateTime) as day,
           COUNT(DISTINCT EmployeeCode) as count
           FROM main_attendance
           WHERE strftime('%Y-%m', AttendanceDateTime) = ?
           GROUP BY day ORDER BY day''',
        (current_month,)
    )

    return jsonify({
        'total': total_emp,
        'present': present_today,
        'absent': max(0, total_emp - present_today),
        'monthly': [{'day': r['day'], 'count': r['count']} for r in monthly_data]
    })


@attendance_bp.route('/employee-history/<path:emp_code>', methods=['GET'])
@login_required
def employee_history(emp_code):
    month = request.args.get('month', date.today().month)
    year = request.args.get('year', date.today().year)
    month = int(month)
    year = int(year)
    days_in_month = calendar.monthrange(year, month)[1]

    rows = query(
        '''SELECT AttendanceDateTime FROM main_attendance
           WHERE EmployeeCode = ?
           AND strftime('%Y', AttendanceDateTime) = ?
           AND strftime('%m', AttendanceDateTime) = ?
           ORDER BY AttendanceDateTime''',
        (emp_code, str(year), str(month).zfill(2))
    )

    by_day = {}
    for r in rows:
        dt_str = r['AttendanceDateTime'] or ''
        day = dt_str[8:10] if len(dt_str) >= 10 else ''
        if not day:
            continue
        by_day.setdefault(day, []).append(dict(r))

    days = []
    for d in range(1, days_in_month + 1):
        key = str(d).zfill(2)
        recs = by_day.get(key, [])
        secs = _calc_hours(recs) if recs else 0
        days.append({
            'day': d,
            'hours': _fmt_hours(secs),
            'seconds': secs,
            'scans': len(recs)
        })

    total_secs = sum(d['seconds'] for d in days)
    present_days = sum(1 for d in days if d['scans'] > 0)

    return jsonify({
        'days': days,
        'total_hours': _fmt_hours(total_secs),
        'present_days': present_days,
        'working_days': days_in_month
    })


@attendance_bp.route('/employee-day-detail/<path:emp_code>', methods=['GET'])
@login_required
def employee_day_detail(emp_code):
    day = request.args.get('date')
    if not day:
        return jsonify({'error': 'date param required'}), 400

    rows = query(
        '''SELECT a.AttendanceDateTime, a.Latitude, a.Longitude,
                  a.Type, a.MarkedBy, a.ProjectId,
                  p.Title AS ProjectName,
                  p.Latitude AS ProjectLat, p.Longitude AS ProjectLon
           FROM main_attendance a
           LEFT JOIN main_myproject p ON a.ProjectId = p.id
           WHERE a.EmployeeCode = ? AND date(a.AttendanceDateTime) = ?
           ORDER BY a.AttendanceDateTime''',
        (emp_code, day)
    )
    scans = []
    for r in rows:
        d = dict(r)
        d['ProjectName'] = d.get('ProjectName') or ''
        scans.append(d)
    return jsonify({'scans': scans})


@attendance_bp.route('/', methods=['POST'])
@login_required
def mark_attendance():
    data = request.get_json()
    emp_code = data.get('EmployeeCode')
    if not emp_code:
        return jsonify({'error': 'EmployeeCode required'}), 400

    emp = query('SELECT * FROM main_employee WHERE EmployeeCode=? AND Deleted=0', (emp_code,), one=True)
    if not emp:
        return jsonify({'error': 'Employee not found'}), 404

    now_dt = datetime.now()
    now = now_dt.strftime('%Y-%m-%d %H:%M:%S')
    att_type = data.get('Type', 'In')
    project_id = data.get('ProjectId')

    auto_signed_out = None
    if att_type == 'In' and project_id:
        today_str = date.today().isoformat()
        open_sessions = query(
            '''SELECT ProjectId, COUNT(*) as cnt
               FROM main_attendance
               WHERE EmployeeCode=? AND date(AttendanceDateTime)=?
               GROUP BY ProjectId''',
            (emp_code, today_str)
        )
        offset = 0
        for s in open_sessions:
            if str(s['ProjectId']) != str(project_id) and s['cnt'] % 2 == 1:
                offset += 1
                auto_out_time = (now_dt - timedelta(seconds=offset)).strftime('%Y-%m-%d %H:%M:%S')
                execute(
                    '''INSERT INTO main_attendance
                       (EmployeeCode, AttendanceDateTime, Latitude, Longitude, MarkedBy, Type, ProjectId)
                       VALUES (?,?,?,?,?,?,?)''',
                    (emp_code, auto_out_time,
                     data.get('Latitude'), data.get('Longitude'),
                     session.get('user_id', ''),
                     'Out', s['ProjectId'])
                )
                auto_signed_out = s['ProjectId']

    execute(
        '''INSERT INTO main_attendance
           (EmployeeCode, AttendanceDateTime, Latitude, Longitude, MarkedBy, Type, ProjectId)
           VALUES (?,?,?,?,?,?,?)''',
        (
            emp_code, now,
            data.get('Latitude'), data.get('Longitude'),
            session.get('user_id', ''),
            att_type,
            project_id
        )
    )

    result = {'success': True, 'employee': emp['FullName'], 'time': now}
    if auto_signed_out:
        result['auto_signed_out_project'] = auto_signed_out
    return jsonify(result), 201


@attendance_bp.route('/sync', methods=['POST'])
@login_required
def sync_attendance():
    """Bulk sync from IndexedDB offline records."""
    records = request.get_json()
    if not isinstance(records, list):
        return jsonify({'error': 'Expected array'}), 400

    synced = 0
    errors = []
    for rec in records:
        try:
            emp_code = rec.get('EmployeeCode')
            att_time = rec.get('AttendanceDateTime')
            if not emp_code or not att_time:
                continue
            existing = query(
                'SELECT id FROM main_attendance WHERE EmployeeCode=? AND AttendanceDateTime=?',
                (emp_code, att_time), one=True
            )
            if not existing:
                execute(
                    '''INSERT INTO main_attendance
                       (EmployeeCode, AttendanceDateTime, Latitude, Longitude, MarkedBy, Type, ProjectId)
                       VALUES (?,?,?,?,?,?,?)''',
                    (
                        emp_code, att_time,
                        rec.get('Latitude'), rec.get('Longitude'),
                        rec.get('MarkedBy', session.get('user_id', '')),
                        rec.get('Type', 'In'),
                        rec.get('ProjectId')
                    )
                )
                synced += 1
        except Exception as e:
            errors.append(str(e))

    return jsonify({'synced': synced, 'errors': errors})


@attendance_bp.route('/export/xlsx', methods=['GET'])
@login_required
def export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    date_filter = request.args.get('date')
    month = request.args.get('month')
    year = request.args.get('year')
    project_id = request.args.get('project_id')

    is_monthly = bool(month and year and not date_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    header_fill = PatternFill(start_color='1a3c5e', end_color='1a3c5e', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

    legend_cell = ws.cell(row=1, column=1, value='🔴 Red highlighted rows/cells indicate attendance marked more than 250m from the project location.')
    legend_cell.font = Font(color='CC0000', italic=True, size=9)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    data_start = 3

    if is_monthly:
        rows, num_days = _build_monthly_summary(month, year, project_id)
        headers = ['#', 'Employee Code', 'Employee Name', 'Designation'] + \
                  [str(d) for d in range(1, num_days + 1)]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=data_start, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        faded_font = Font(color='C0C0C0')
        for i, r in enumerate(rows, 1):
            row_num = data_start + i
            row_data = [i, r['code'], r['name'], r['designation']] + \
                       [r['days'].get(d, '') for d in range(1, num_days + 1)]
            for col, val in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=val)
            for d in range(1, num_days + 1):
                col_idx = 4 + d
                cell = ws.cell(row=row_num, column=col_idx)
                if d in r.get('flagged_days', set()):
                    cell.fill = red_fill
                elif cell.value == '0':
                    cell.font = faded_font
    else:
        summary = _build_export_summary(date_filter, month, year, project_id)
        headers = ['#', 'Employee Code', 'Employee Name', 'Designation',
                   'First In', 'Last Out', 'Total Hours', 'Status']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=data_start, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for i, r in enumerate(summary, 1):
            row_num = data_start + i
            row_data = [i, r['code'], r['name'], r['designation'],
                        r['first_in'], r['last_out'], r['hours'], r['status']]
            for col, val in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=val)
            if r.get('out_of_range'):
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col_idx).fill = red_fill

    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        cells = [c for c in col if not isinstance(c, MergedCell)]
        if not cells:
            continue
        max_len = max((len(str(c.value or '')) for c in cells), default=10)
        ws.column_dimensions[cells[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"attendance_{date_filter or f'{year}-{month}'}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@attendance_bp.route('/export/pdf', methods=['GET'])
@login_required
def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    date_filter = request.args.get('date')
    month = request.args.get('month')
    year = request.args.get('year')
    project_id = request.args.get('project_id')

    is_monthly = bool(month and year and not date_filter)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    title = f"Attendance Report — {date_filter or f'{year}/{month}'}"
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 0.4*cm))

    red_bg = colors.HexColor('#FFCCCC')

    if is_monthly:
        rows, num_days = _build_monthly_summary(month, year, project_id)
        data = [['#', 'Code', 'Name', 'Desig.'] + [str(d) for d in range(1, num_days + 1)]]
        for i, r in enumerate(rows, 1):
            data.append(
                [i, r['code'], r['name'], r['designation']] +
                [r['days'].get(d, '') for d in range(1, num_days + 1)]
            )
        font_size = 5 if num_days > 28 else 6
    else:
        summary = _build_export_summary(date_filter, month, year, project_id)
        data = [['#', 'Code', 'Name', 'Designation', 'First In', 'Last Out', 'Hours', 'Status']]
        for i, r in enumerate(summary, 1):
            data.append([
                i, r['code'], r['name'], r['designation'],
                r['first_in'], r['last_out'], r['hours'], r['status']
            ])
        font_size = 8

    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c5e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), font_size),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 3),
    ]
    if is_monthly:
        faded = colors.HexColor('#c0c0c0')
        flagged_days_list = [r.get('flagged_days', set()) for r in rows]
        for row_idx, row in enumerate(data[1:], 1):
            flagged = flagged_days_list[row_idx - 1]
            for col_idx in range(4, len(row)):
                d = col_idx - 3
                if d in flagged:
                    style_cmds.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), red_bg))
                elif row[col_idx] == '0':
                    style_cmds.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), faded))
    else:
        for row_idx, r in enumerate(summary, 1):
            if r.get('out_of_range'):
                style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), red_bg))

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    filename = f"attendance_{date_filter or f'{year}-{month}'}.pdf"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


def _format_time(dt_str):
    """Convert 'YYYY-MM-DD HH:MM:SS' to '09:30 AM' local format."""
    if not dt_str or len(dt_str) < 16:
        return '—'
    try:
        t = datetime.strptime(dt_str.strip(), '%Y-%m-%d %H:%M:%S')
        return t.strftime('%I:%M %p')
    except ValueError:
        return dt_str[11:16]


def _calc_hours(records):
    """Pair In/Out scans and return total seconds worked."""
    total = 0
    i = 0
    while i < len(records) - 1:
        in_time = records[i].get('AttendanceDateTime', '')
        out_time = records[i + 1].get('AttendanceDateTime', '')
        if in_time and out_time:
            try:
                t_in = datetime.strptime(in_time, '%Y-%m-%d %H:%M:%S')
                t_out = datetime.strptime(out_time, '%Y-%m-%d %H:%M:%S')
                diff = (t_out - t_in).total_seconds()
                if diff > 0:
                    total += diff
            except ValueError:
                pass
        i += 2
    return total


def _fmt_hours(total_seconds):
    """Format seconds as 'H:MM'."""
    hrs = int(total_seconds // 3600)
    mins = int((total_seconds % 3600) // 60)
    return f'{hrs}:{mins:02d}'


def _haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _get_project_locations():
    rows = query('SELECT id, Latitude, Longitude FROM main_myproject WHERE Deleted=0')
    return {
        r['id']: (float(r['Latitude']), float(r['Longitude']))
        for r in rows if r['Latitude'] and r['Longitude']
    }


def _is_out_of_range(record, proj_locs):
    att_lat = record.get('Latitude')
    att_lon = record.get('Longitude')
    proj_id = record.get('ProjectId')
    if not att_lat or not att_lon or not proj_id:
        return False
    proj_loc = proj_locs.get(int(proj_id) if isinstance(proj_id, str) else proj_id)
    if not proj_loc:
        return False
    try:
        return _haversine(float(att_lat), float(att_lon), proj_loc[0], proj_loc[1]) > 250
    except (ValueError, TypeError):
        return False


def _build_monthly_summary(month, year, project_id):
    """Build per-employee monthly grid with daily hours."""
    employees = query(
        'SELECT EmployeeCode, FullName, Designation FROM main_employee '
        'WHERE Deleted=0 ORDER BY FullName'
    )
    att_rows = _fetch_attendance(None, month, year, project_id)
    proj_locs = _get_project_locations()
    num_days = calendar.monthrange(int(year), int(month))[1]

    by_emp_day = {}
    for r in att_rows:
        code = r['EmployeeCode']
        dt_str = r.get('AttendanceDateTime', '')
        if len(dt_str) < 10:
            continue
        day = int(dt_str[8:10])
        key = (code, day)
        if key not in by_emp_day:
            by_emp_day[key] = []
        by_emp_day[key].append(r)

    summary = []
    for emp in employees:
        code = emp['EmployeeCode']
        row = {
            'code': code,
            'name': emp['FullName'],
            'designation': emp['Designation'] or '',
            'days': {},
            'flagged_days': set()
        }
        for d in range(1, num_days + 1):
            records = by_emp_day.get((code, d), [])
            records.sort(key=lambda r: r.get('AttendanceDateTime', ''))
            secs = _calc_hours(records)
            row['days'][d] = _fmt_hours(secs) if secs > 0 else '0'
            if any(_is_out_of_range(rec, proj_locs) for rec in records):
                row['flagged_days'].add(d)
        summary.append(row)

    return summary, num_days


def _build_export_summary(date_filter, month, year, project_id):
    """Build per-employee summary with total hours worked."""
    employees = query(
        'SELECT EmployeeCode, FullName, Designation FROM main_employee '
        'WHERE Deleted=0 ORDER BY FullName'
    )

    att_rows = _fetch_attendance(date_filter, month, year, project_id)
    proj_locs = _get_project_locations()

    by_emp = OrderedDict()
    for r in att_rows:
        code = r['EmployeeCode']
        if code not in by_emp:
            by_emp[code] = []
        by_emp[code].append(r)

    summary = []
    for emp in employees:
        code = emp['EmployeeCode']
        records = by_emp.get(code, [])
        records.sort(key=lambda r: r.get('AttendanceDateTime', ''))

        if not records:
            summary.append({
                'code': code,
                'name': emp['FullName'],
                'designation': emp['Designation'] or '',
                'first_in': '—',
                'last_out': '—',
                'hours': '0:00',
                'status': 'Absent',
                'out_of_range': False
            })
            continue

        first_in = _format_time(records[0].get('AttendanceDateTime', '')) if records else '—'
        last_out = '—'
        if len(records) >= 2 and len(records) % 2 == 0:
            last_out = _format_time(records[-1].get('AttendanceDateTime', ''))

        total_seconds = _calc_hours(records)
        still_in = len(records) % 2 == 1
        flagged = any(_is_out_of_range(rec, proj_locs) for rec in records)
        summary.append({
            'code': code,
            'name': emp['FullName'],
            'designation': emp['Designation'] or '',
            'first_in': first_in,
            'last_out': last_out if not still_in else '—',
            'hours': _fmt_hours(total_seconds),
            'status': 'Still In' if still_in else 'Present',
            'out_of_range': flagged
        })

    return summary


def _fetch_attendance(date_filter, month, year, project_id):
    """Fetch raw attendance records for the given period."""
    if month and year:
        sql = '''SELECT a.EmployeeCode, a.AttendanceDateTime,
                        a.Latitude, a.Longitude, a.ProjectId
                 FROM main_attendance a
                 WHERE strftime('%m', a.AttendanceDateTime) = ?
                 AND strftime('%Y', a.AttendanceDateTime) = ?
                 ORDER BY a.EmployeeCode, a.AttendanceDateTime'''
        params = (month.zfill(2), str(year))
    elif date_filter:
        sql = '''SELECT a.EmployeeCode, a.AttendanceDateTime,
                        a.Latitude, a.Longitude, a.ProjectId
                 FROM main_attendance a
                 WHERE date(a.AttendanceDateTime) = ?
                 ORDER BY a.EmployeeCode, a.AttendanceDateTime'''
        params = (date_filter,)
    else:
        sql = '''SELECT a.EmployeeCode, a.AttendanceDateTime,
                        a.Latitude, a.Longitude, a.ProjectId
                 FROM main_attendance a
                 WHERE date(a.AttendanceDateTime) = ?
                 ORDER BY a.EmployeeCode, a.AttendanceDateTime'''
        params = (date.today().isoformat(),)

    if project_id:
        sql = sql.replace('ORDER BY', 'AND a.ProjectId = ? ORDER BY')
        params = params + (project_id,)

    return [dict(r) for r in query(sql, params)]
